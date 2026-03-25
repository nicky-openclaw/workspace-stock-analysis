#!/usr/bin/env python3
"""
板块效应分析 v4.3
数据获取优先级：
1. 腾讯API → 个股涨跌幅（免费）
2. mx_finance_data → 行业归属+板块涨跌幅（一次自然语言查询搞定）
3. eastmoney_financial_data → 主力净额（JSON稳定解析）
4. agent-browser → 补齐行业（两者都失败时）

v4.3优化：
- mx_finance_data 作为行业+板块涨跌的首选（一次自然语言搞定多字段）
- eastmoney_financial_data 专攻主力净额（稳定JSON解析）
- 减少API调用次数（从2次/股→1次行业涨跌+1次主力）
"""

import json
import subprocess
import os
import sys
import argparse
import urllib.request
import asyncio
import openpyxl
from pathlib import Path
from datetime import datetime

WORKSPACE = Path("/Users/nicky/.openclaw/workspace-stock-analysis")
MX_SCRIPT = Path.home() / ".openclaw/workspace/skills/mx-finance-data/scripts/get_data.py"
EASTMONEY_APIKEY = "mkt_ed_FmsusuPQr6aZCpqc2Pgof6l7gGbnvS_riNSxtGeI"
EASTMONEY_DATA_URL = "https://mkapi2.dfcfs.com/finskillshub/api/claw/query"


# ══════════════════════════════════════════════════════════════════════════════
# 工具函数
# ══════════════════════════════════════════════════════════════════════════════

def get_stock_tencent(codes: list) -> dict:
    """腾讯API获取个股涨跌幅"""
    result = {}
    try:
        url = f"https://qt.gtimg.cn/q={','.join(codes)}"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        resp = urllib.request.urlopen(req, timeout=10)
        data = resp.read().decode('gbk', errors='ignore')
        for line in data.split(';'):
            if '~' not in line:
                continue
            if 'v_' in line:
                code = line.split('v_')[1].split('=')[0]
            else:
                continue
            parts = line.split('~')
            name = parts[1] if len(parts) > 1 else ""
            change = float(parts[32]) if len(parts) > 32 and parts[32] else 0
            result[code] = {"name": name, "change_pct": change}
    except Exception as e:
        print(f"  ⚠️ 腾讯API失败: {e}")
    return result


def get_industry_browser(code: str) -> str:
    """用agent-browser获取个股行业归属（最终保底）"""
    try:
        if code.startswith('6') or code.startswith('688'):
            url = f"https://quote.eastmoney.com/sh{code}.html"
        else:
            url = f"https://quote.eastmoney.com/sz{code}.html"
        r1 = subprocess.run(["agent-browser", "open", url],
                           capture_output=True, text=True, timeout=15)
        subprocess.run(["agent-browser", "wait", "2000"],
                      capture_output=True, text=True, timeout=5)
        r2 = subprocess.run(["agent-browser", "get", "text", "@e89"],
                           capture_output=True, text=True, timeout=10)
        industry = r2.stdout.strip()
        if industry and len(industry) < 20:
            return industry
    except Exception as e:
        pass
    return ""


# ══════════════════════════════════════════════════════════════════════════════
# mx_finance_data（自然语言查询，行业+板块涨跌一次搞定）
# ══════════════════════════════════════════════════════════════════════════════

def get_industry_sector_from_mx(stock_code: str, stock_name: str) -> dict:
    """
    用mx_finance_data自然语言查询行业归属+板块涨跌幅。
    一次查询返回行业（申万）+ 板块涨跌，返回xlsx需openpyxl解析。
    成功返回 {"industry": "电子", "sector_change_pct": "2.65%"}，失败返回空dict。
    """
    try:
        query = (
            f"{stock_name}({stock_code})属于什么行业？"
            f"今日这个行业的涨跌幅是多少？"
        )
        result = subprocess.run(
            ["python3", str(MX_SCRIPT), "--query", query],
            capture_output=True, text=True, timeout=60,
            env={**os.environ, "EM_API_KEY": os.environ.get("EM_API_KEY", "")}
        )
        if result.returncode != 0:
            return {}

        # 解析xlsx文件路径（从stdout最后两行提取）
        stdout = result.stdout.strip()
        lines = [l for l in stdout.split('\n') if l.startswith('文件:')]
        if not lines:
            return {}
        xlsx_path = lines[-1].replace('文件:', '').strip()

        if not Path(xlsx_path).exists():
            return {}

        # 解析xlsx：取第一个sheet的数据
        wb = openpyxl.load_workbook(xlsx_path)
        industry = ""
        sector_change = ""

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            # 扫描所有行，匹配行业关键词
            for row in rows:
                row_str = ' '.join([str(c) if c else '' for c in row])
                # 找行业（申万行业）
                if ('申万' in row_str or '行业分类' in row_str) and not industry:
                    for cell in row:
                        if cell and ('行业' in str(cell) or '申万' in str(cell)):
                            ind = str(cell).replace('(申万)', '').strip()
                            if ind and len(ind) < 15:
                                industry = ind
                # 找涨跌幅
                if ('涨跌幅' in row_str or '%' in row_str) and not sector_change:
                    for i, cell in enumerate(row):
                        val_str = str(cell) if cell else ''
                        if val_str.replace('.', '').replace('-', '').isdigit() or (val_str.endswith('%') and i > 0):
                            try:
                                if val_str.endswith('%'):
                                    sector_change = val_str
                                elif '.' in val_str:
                                    sector_change = f"{float(val_str):.2f}%"
                            except:
                                pass

        if industry or sector_change:
            return {
                "industry": industry,
                "sector_change_pct": sector_change
            }
        return {}

    except Exception as e:
        return {}


# ══════════════════════════════════════════════════════════════════════════════
# eastmoney_financial_data（JSON稳定解析，主力净额专用）
# ══════════════════════════════════════════════════════════════════════════════

def get_industry_from_eastmoney(stock_code: str, stock_name: str) -> dict:
    """用eastmoney_financial_data获取个股所属行业（申万行业）"""
    try:
        query = f"{stock_name}({stock_code})所属申万行业指数名称"
        payload = json.dumps({"toolQuery": query}).encode()
        req = urllib.request.Request(
            EASTMONEY_DATA_URL,
            data=payload,
            headers={
                'Content-Type': 'application/json',
                'apikey': EASTMONEY_APIKEY
            },
            method='POST'
        )
        resp = urllib.request.urlopen(req, timeout=15)
        d = json.loads(resp.read())
        tables = d.get('data',{}).get('data',{}).get('searchDataResultDTO',{}).get('dataTableDTOList',[])
        if not tables:
            return {}
        t = tables[0]
        raw = t.get('rawTable', {}) or t.get('table', {})
        industry = ""
        for k, v in raw.items():
            if k != 'headName' and isinstance(v, list) and v:
                industry = v[0].replace('(申万)', '').strip()
                break
        return {'industry': industry}
    except Exception as e:
        return {}


def get_main_inflow_from_eastmoney(industry: str) -> dict:
    """
    用eastmoney_financial_data获取指定行业板块的主力净额（f62字段，稳定JSON）。
    返回 {"sector_main_inflow": "xxx"}，失败返回空dict。
    """
    if not industry:
        return {}
    try:
        query = f"{industry}行业今日主力净流入多少亿？"
        payload = json.dumps({"toolQuery": query}).encode()
        req = urllib.request.Request(
            EASTMONEY_DATA_URL,
            data=payload,
            headers={
                'Content-Type': 'application/json',
                'apikey': EASTMONEY_APIKEY
            },
            method='POST'
        )
        resp = urllib.request.urlopen(req, timeout=15)
        d = json.loads(resp.read())
        tables = d.get('data',{}).get('data',{}).get('searchDataResultDTO',{}).get('dataTableDTOList',[])
        if not tables:
            return {}
        t = tables[0]
        raw = t.get('rawTable', {}) or t.get('table', {})
        f62 = raw.get('f62', [None])[0] if isinstance(raw.get('f62'), list) else raw.get('f62')
        return {'sector_main_inflow': f62}
    except Exception as e:
        return {}


# ══════════════════════════════════════════════════════════════════════════════
# 格式化
# ══════════════════════════════════════════════════════════════════════════════

def format_money(val):
    if val is None or val == 'N/A':
        return "N/A"
    try:
        s = str(val)
        if '亿' in s or '万' in s:
            return s.replace('元','')
        f = float(s)
        if abs(f) >= 1e8: return f"{f/1e8:+.2f}亿"
        elif abs(f) >= 1e4: return f"{f/1e4:+.2f}万"
        return f"{f:+.0f}"
    except:
        return str(val)


def format_pct(val):
    if val is None or val == 'N/A':
        return "N/A"
    s = str(val)
    if '%' in s:
        return s
    try:
        return f"{float(s):+.2f}%"
    except:
        return s


# ══════════════════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════════════════

def analyze(framework: str, top_n: int = 10):
    today = datetime.now().strftime("%Y-%m-%d")
    scores_file = WORKSPACE / f"{framework}_output" / f"{framework}_scores.json"
    if not scores_file.exists():
        print(f"❌ 找不到 {scores_file}"); sys.exit(1)

    with open(scores_file) as f:
        data = json.load(f)

    all_stocks = []
    if "top_limit_up" in data:
        all_stocks = data.get("top_limit_up", [])[:5] + data.get("top_non_limit_up", [])[:5]
    elif "top5" in data:
        all_stocks = data.get("top5", [])[:top_n]
    else:
        all_stocks = data.get("all_scored", [])[:top_n]

    print(f"\n{'='*70}")
    print(f"📊 板块效应分析 - {framework.upper()} | {today}")
    print(f"{'='*70}")
    print(f"分析股票数: {len(all_stocks)}")

    # Step 1: 腾讯API获取涨跌幅
    print("\n⏳ Step 1: 腾讯API获取涨跌幅...")
    codes = [s["code"] for s in all_stocks]
    tencent_codes = []
    for c in codes:
        if c.startswith('6') or c.startswith('688'):
            tencent_codes.append('sh' + c)
        else:
            tencent_codes.append('sz' + c)
    stock_data = get_stock_tencent(tencent_codes)
    print(f"  ✅ 获取 {len(stock_data)} 只股票涨跌幅")

    # 读取缓存
    cache_file = WORKSPACE / f"{framework}_output" / "sector_cache.json"
    sector_cache = {}
    if cache_file.exists():
        with open(cache_file) as f:
            sector_cache = json.load(f)

    # Step 2: mx_finance_data 获取行业+板块涨跌（自然语言，一次搞定）
    print("\n⏳ Step 2: mx_finance_data获取行业+板块涨跌...")
    mx_count = 0
    for s in all_stocks:
        code, name = s["code"], s["name"]
        # 已有完整数据则跳过
        c = sector_cache.get(code, {})
        if c.get("industry") and c.get("sector_change_pct"):
            continue
        # 优先用mx自然语言获取
        if not c.get("industry") or not c.get("sector_change_pct"):
            result = get_industry_sector_from_mx(code, name)
            if result:
                sector_cache[code] = {**c, **result}
                mx_count += 1
                ind = result.get('industry', 'N/A')
                chg = result.get('sector_change_pct', 'N/A')
                print(f"  ✅ {name}: 行业={ind}, 涨跌={chg}")
    print(f"  共获取 {mx_count} 条数据")

    # Step 3: eastmoney_financial_data 补充主力净额
    print("\n⏳ Step 3: eastmoney_financial_data获取主力净额...")
    em_count = 0
    for s in all_stocks:
        code, name = s["code"], s["name"]
        c = sector_cache.get(code, {})
        industry = c.get("industry", "")
        # 有行业但缺主力净额
        if industry and not c.get("sector_main_inflow"):
            inflow_result = get_main_inflow_from_eastmoney(industry)
            if inflow_result:
                sector_cache[code] = {**c, **inflow_result}
                em_count += 1
                print(f"  ✅ {name}({industry}): 主力={inflow_result.get('sector_main_inflow', 'N/A')}")
    print(f"  共获取 {em_count} 条数据")

    # Step 4: eastmoney_financial_data 补充缺失的行业+涨跌
    print("\n⏳ Step 4: eastmoney_financial_data补齐行业和板块涨跌...")
    em_count2 = 0
    for s in all_stocks:
        code, name = s["code"], s["name"]
        c = sector_cache.get(code, {})
        # 缺行业
        if not c.get("industry"):
            ind_result = get_industry_from_eastmoney(code, name)
            if ind_result.get('industry'):
                sector_cache[code] = {**c, **ind_result}
                em_count2 += 1
                print(f"  ✅ {name}: 行业={ind_result.get('industry','')}")
        # 有行业但缺板块涨跌
        if sector_cache.get(code, {}).get("industry") and not sector_cache[code].get("sector_change_pct"):
            industry = sector_cache[code]["industry"]
            # 查板块涨跌
            try:
                query = f"{industry}行业今日涨跌幅"
                payload = json.dumps({"toolQuery": query}).encode()
                req = urllib.request.Request(
                    EASTMONEY_DATA_URL,
                    data=payload,
                    headers={'Content-Type': 'application/json', 'apikey': EASTMONEY_APIKEY},
                    method='POST'
                )
                resp = urllib.request.urlopen(req, timeout=15)
                d = json.loads(resp.read())
                tables = d.get('data',{}).get('data',{}).get('searchDataResultDTO',{}).get('dataTableDTOList',[])
                if tables:
                    raw = tables[0].get('rawTable', {}) or tables[0].get('table', {})
                    f3 = raw.get('f3', [None])[0] if isinstance(raw.get('f3'), list) else raw.get('f3')
                    if f3:
                        f3_str = f3 if str(f3).endswith('%') else f"{float(f3):.2f}%"
                        sector_cache[code]["sector_change_pct"] = f3_str
                        em_count2 += 1
                        print(f"  ✅ {name}: 涨跌={f3_str}")
            except:
                pass
    if em_count2:
        print(f"  共补齐 {em_count2} 条数据")

    # Step 5: agent-browser 最终保底行业
    print("\n⏳ Step 5: agent-browser补齐缺失行业...")
    browser_count = 0
    for s in all_stocks:
        code = s["code"]
        if not sector_cache.get(code, {}).get("industry"):
            industry = get_industry_browser(code)
            if industry:
                sector_cache[code] = sector_cache.get(code, {})
                sector_cache[code]["industry"] = industry
                browser_count += 1
                print(f"  ✅ {s['name']}({code}): {industry}")
    if browser_count:
        print(f"  共补齐 {browser_count} 只股票行业")
    else:
        print(f"  无需补齐")

    # 保存缓存
    with open(cache_file, "w") as f:
        json.dump(sector_cache, f, ensure_ascii=False, indent=2)

    # 输出结果
    print(f"\n{'─'*80}")
    print(f"{'代码':<8} {'名称':<10} {'涨跌%':>6} {'所属板块':<14} {'板块涨跌':>8} {'板块主力':>12}")
    print(f"{'─'*80}")
    for s in all_stocks:
        code, name = s["code"], s["name"]
        t = stock_data.get('sh'+code if code.startswith('6') else 'sz'+code, {})
        c = sector_cache.get(code, {})
        change = t.get("change_pct", s.get("change_pct", 0))
        industry = c.get("industry", "⚠️待获取")
        sector_change = c.get("sector_change_pct", "N/A")
        sector_inflow = c.get("sector_main_inflow", "N/A")
        print(f"{code:<8} {name:<10} {change:>+6.2f}% {industry:<14} {format_pct(sector_change):>8} {format_money(sector_inflow):>12}")

    total_calls = mx_count + em_count + em_count2
    print(f"\n✅ 分析完成！")
    print(f"  腾讯API: 获取涨跌幅")
    print(f"  mx_finance_data: 行业+板块涨跌 {mx_count}次")
    print(f"  eastmoney_financial_data: 主力净额 {em_count}次 + 补齐 {em_count2}次")
    print(f"  agent-browser: 补齐行业 {browser_count}只")


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--framework", choices=["ztx", "qdk", "b1"], default="ztx")
    p.add_argument("--top", type=int, default=10)
    args = p.parse_args()
    analyze(args.framework, args.top)
